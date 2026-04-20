import os
import re
import io
import csv
import logging
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional, Dict

import vobject
import openpyxl
import phonenumbers

try:
    import pytesseract
    from PIL import Image
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False

from telegram import (
    Update, InputFile, Message,
    ReplyKeyboardMarkup, ReplyKeyboardRemove,
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ConversationHandler,
    ContextTypes, filters,
)

# =============================================================================
# CONFIGURATION
# =============================================================================

BOT_TOKEN = "8247588556:AAGgTKtILWZL6mTQEvLAziJmrbAkRTUPA0E"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# Default user settings
DEFAULT_SETTINGS = {
    "default_format": "vcf",
    "per_file": 500,
    "contact_name": "Contact",
    "start_number": 1,
}

# =============================================================================
# MENU BUTTON LABELS
# =============================================================================

BTN_FA    = "📊 File Analysis"
BTN_FC    = "🔄 File Converter"
BTN_QV    = "⚡ Quick VCF"
BTN_VM    = "🛠 VCF Maker"
BTN_SF    = "✂️ Split File"
BTN_MF    = "🔗 Merge Files"
BTN_FE    = "✏️ File Editor"
BTN_LM    = "📋 List Maker"
BTN_RF    = "📝 Rename File"
BTN_RC    = "👤 Rename Contact"
BTN_ST    = "⚙️ Settings"
BTN_RS    = "🔄 Reset"
BTN_HELP  = "❓ Help"
BTN_BACK  = "🔙 Back"
BTN_CANCEL= "❌ Cancel"
BTN_DONE  = "✅ Done"

MENU_PATTERN = (
    r"^(📊 File Analysis|🔄 File Converter|⚡ Quick VCF|🛠 VCF Maker|"
    r"✂️ Split File|🔗 Merge Files|✏️ File Editor|📋 List Maker|"
    r"📝 Rename File|👤 Rename Contact|⚙️ Settings|🔄 Reset|❓ Help)$"
)

# =============================================================================
# KEYBOARD BUILDERS
# =============================================================================

def main_menu_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [BTN_FA,   BTN_FC],
        [BTN_QV,   BTN_VM],
        [BTN_SF,   BTN_MF],
        [BTN_FE,   BTN_LM],
        [BTN_RF,   BTN_RC],
        [BTN_ST,   BTN_RS],
        [BTN_HELP],
    ], resize_keyboard=True)


def back_cancel_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_BACK, BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
    )


def done_cancel_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[BTN_DONE, BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
    )


def format_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        ["📄 TXT", "📇 VCF"],
        ["📊 CSV", "📑 XLSX"],
        [BTN_BACK, BTN_CANCEL],
    ], resize_keyboard=True, one_time_keyboard=True)


def yes_no_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [["✅ Yes, Reset", BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
    )


def more_finish_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [["➕ Add More", "✅ Finish"], [BTN_CANCEL]],
        resize_keyboard=True, one_time_keyboard=True
    )


def skip_cancel_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [["⏭ Skip", BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
    )


def confirm_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [["✅ Generate", BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
    )


def rename_mode_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [["👥 Rename ALL", "👤 Rename SINGLE"], [BTN_CANCEL]],
        resize_keyboard=True, one_time_keyboard=True
    )


def settings_kb(settings: dict) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [f"📄 Format: {settings['default_format'].upper()}"],
        [f"📦 Per File: {settings['per_file']}"],
        [f"👤 Contact Name: {settings['contact_name']}"],
        [f"🔢 Start Number: {settings['start_number']}"],
        [BTN_BACK],
    ], resize_keyboard=True)


def editor_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        ["◀️ Prev", "▶️ Next"],
        ["✏️ Edit", "🗑 Remove"],
        ["➕ Add", "💾 Save"],
        [BTN_BACK],
    ], resize_keyboard=True)

# =============================================================================
# CONVERSATION STATES
# =============================================================================

# File Analysis
FA_UPLOAD = 0

# File Converter
FC_UPLOAD, FC_FORMAT = 0, 1

# Quick VCF
QV_FILENAME, QV_NAME, QV_PHONE, QV_MORE = 0, 1, 2, 3

# VCF Maker — new 9-state flow (CHANGE 3)
VM_UPLOAD, VM_BASENAME, VM_CONTACT_NAME, VM_PER_FILE, VM_CONTACT_START, VM_FILE_START, VM_GROUP_NAME, VM_GROUP_START, VM_CONFIRM = range(9)

# Split File
SF_UPLOAD, SF_COUNT = 0, 1

# Merge Files
MF_UPLOAD = 0

# File Editor
FE_UPLOAD, FE_VIEW, FE_EDIT_SELECT, FE_EDIT_NAME, FE_EDIT_PHONE, FE_REMOVE_IDX, FE_ADD_NAME, FE_ADD_PHONE = range(8)

# List Maker
LM_UPLOAD = 0

# Rename File
RF_UPLOAD, RF_NAME = 0, 1

# Rename Contact
RC_UPLOAD, RC_MODE, RC_ALL_NAME, RC_SINGLE_SELECT, RC_SINGLE_NAME = range(5)

# Settings
ST_MAIN, ST_SET_PER_FILE, ST_SET_CONTACT_NAME, ST_SET_START_NUM = range(4)

# Reset
RS_CONFIRM = 0

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_settings(context: ContextTypes.DEFAULT_TYPE) -> dict:
    if "settings" not in context.user_data:
        context.user_data["settings"] = DEFAULT_SETTINGS.copy()
    for k, v in DEFAULT_SETTINGS.items():
        context.user_data["settings"].setdefault(k, v)
    return context.user_data["settings"]


# CHANGE 1: Keep letters, digits, hyphens, underscores. Remove everything else.
def clean_name(name: str) -> str:
    """Keep letters, digits, hyphens, underscores. Remove everything else."""
    return re.sub(r'[^a-zA-Z0-9_\-]', '', name)


# CHANGE 2: ALWAYS add + prefix to every phone number.
def clean_phone(raw: str) -> str:
    raw = raw.strip()
    digits = re.sub(r'\D', '', raw)
    if len(digits) < 7:
        return ""
    return '+' + digits


def parse_vcf_contacts(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    try:
        text = data.decode("utf-8", errors="replace")
        for vcard in vobject.readComponents(text):
            name = ""
            phone = ""
            try:
                name = str(vcard.fn.value).strip()
            except Exception:
                pass
            try:
                tel = vcard.tel.value
                phone = clean_phone(str(tel))
            except Exception:
                pass
            if not phone:
                try:
                    for tel in vcard.contents.get("tel", []):
                        p = clean_phone(str(tel.value))
                        if p:
                            phone = p
                            break
                except Exception:
                    pass
            if phone:
                contacts.append((name or "Contact", phone))
    except Exception as e:
        logger.error(f"VCF parse error: {e}")
    return contacts


def parse_txt_contacts(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    text = data.decode("utf-8", errors="replace")
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if ',' in line:
            parts = line.split(',', 1)
            name = parts[0].strip()
            phone = clean_phone(parts[1].strip())
            if phone:
                contacts.append((name or "Contact", phone))
        else:
            phone = clean_phone(line)
            if phone:
                contacts.append(("Contact", phone))
    return contacts


def parse_csv_contacts(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if not row:
            continue
        if len(row) >= 2:
            name = row[0].strip()
            phone = clean_phone(row[1].strip())
            if phone:
                contacts.append((name or "Contact", phone))
        elif len(row) == 1:
            phone = clean_phone(row[0].strip())
            if phone:
                contacts.append(("Contact", phone))
    return contacts


def parse_xlsx_contacts(data: bytes) -> List[Tuple[str, str]]:
    contacts = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            if len(row) >= 2 and row[1] is not None:
                name = str(row[0]).strip() if row[0] else "Contact"
                phone = clean_phone(str(row[1]).strip())
                if phone:
                    contacts.append((name, phone))
            elif row[0] is not None:
                phone = clean_phone(str(row[0]).strip())
                if phone:
                    contacts.append(("Contact", phone))
    except Exception as e:
        logger.error(f"XLSX parse error: {e}")
    return contacts


def parse_file(data: bytes, filename: str) -> List[Tuple[str, str]]:
    ext = Path(filename).suffix.lower()
    if ext == ".vcf":
        return parse_vcf_contacts(data)
    elif ext == ".txt":
        return parse_txt_contacts(data)
    elif ext == ".csv":
        return parse_csv_contacts(data)
    elif ext in (".xlsx", ".xls"):
        return parse_xlsx_contacts(data)
    else:
        return parse_txt_contacts(data)


def contacts_to_vcf(contacts: List[Tuple[str, str]]) -> bytes:
    lines = []
    for name, phone in contacts:
        lines.append("BEGIN:VCARD")
        lines.append("VERSION:3.0")
        lines.append(f"FN:{name}")
        lines.append(f"TEL;TYPE=CELL:{phone}")
        lines.append("END:VCARD")
        lines.append("")
    return "\n".join(lines).encode("utf-8")


def contacts_to_txt(contacts: List[Tuple[str, str]]) -> bytes:
    return "\n".join(phone for _, phone in contacts).encode("utf-8")


def contacts_to_csv(contacts: List[Tuple[str, str]]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["Name", "Phone"])
    for name, phone in contacts:
        writer.writerow([name, phone])
    return out.getvalue().encode("utf-8")


def contacts_to_xlsx(contacts: List[Tuple[str, str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Phone"])
    for name, phone in contacts:
        ws.append([name, phone])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def contacts_to_format(contacts: List[Tuple[str, str]], fmt: str) -> bytes:
    fmt = fmt.lower()
    if fmt == "vcf":
        return contacts_to_vcf(contacts)
    elif fmt == "txt":
        return contacts_to_txt(contacts)
    elif fmt == "csv":
        return contacts_to_csv(contacts)
    elif fmt == "xlsx":
        return contacts_to_xlsx(contacts)
    return contacts_to_vcf(contacts)


def get_country_from_phone(phone: str) -> str:
    try:
        p = phonenumbers.parse(phone if phone.startswith('+') else '+' + phone)
        region = phonenumbers.region_code_for_number(p)
        return region or "Unknown"
    except Exception:
        return "Unknown"


def analyze_contacts(contacts: List[Tuple[str, str]]) -> dict:
    seen_digits = set()
    clean_count = 0
    duplicate_count = 0
    junk_count = 0
    country_counts: Dict[str, int] = {}

    for name, phone in contacts:
        digits = re.sub(r'\D', '', phone)
        if len(digits) < 7:
            junk_count += 1
            continue
        if digits in seen_digits:
            duplicate_count += 1
        else:
            seen_digits.add(digits)
            clean_count += 1
            country = get_country_from_phone(phone)
            country_counts[country] = country_counts.get(country, 0) + 1

    return {
        "total": len(contacts),
        "clean": clean_count,
        "duplicate": duplicate_count,
        "junk": junk_count,
        "countries": country_counts,
    }


def paginate_contacts(contacts: List[Tuple[str, str]], page: int, per_page: int = 10) -> Tuple[str, int, int]:
    total = len(contacts)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(0, min(page, total_pages - 1))
    start = page * per_page
    end = start + per_page
    lines = [f"📋 *Contacts* (Page {page+1}/{total_pages}):\n"]
    for i, (name, phone) in enumerate(contacts[start:end], start=start+1):
        lines.append(f"`{i}.` {name} — `{phone}`")
    return "\n".join(lines), page, total_pages


def label_to_fmt(label: str) -> str:
    """Convert button label like '📄 TXT' to format string 'txt'."""
    mapping = {
        "📄 TXT": "txt",
        "📇 VCF": "vcf",
        "📊 CSV": "csv",
        "📑 XLSX": "xlsx",
    }
    return mapping.get(label, "vcf")

# =============================================================================
# COMMON HANDLERS
# =============================================================================

async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str = None):
    if text is None:
        text = "Choose a feature from the menu below:"
    msg = update.effective_message
    await msg.reply_text(text, reply_markup=main_menu_kb())


async def restart_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle /start from anywhere — reset to main menu."""
    user = update.effective_user
    now = datetime.now()
    text = (
        f"👋 Welcome, {user.full_name}!\n\n"
        f"📋 Your Profile:\n"
        f"┌ 👤 Name: {user.full_name}\n"
        f"├ 🆔 ID: {user.id}\n"
        f"├ 📛 Username: @{user.username or 'N/A'}\n"
        f"├ 📅 Date: {now.strftime('%d %B %Y')}\n"
        f"└ 🕐 Time: {now.strftime('%I:%M %p')}\n\n"
        f"🤖 VCF Contact Bot\n"
        f"Choose a feature from the menu below!"
    )
    await update.effective_message.reply_text(text, reply_markup=main_menu_kb())
    return ConversationHandler.END


async def cancel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Universal cancel — show main menu."""
    await show_main_menu(update, context, "❌ Cancelled. Choose a feature:")
    return ConversationHandler.END


async def menu_redirect(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Redirect mid-conversation menu button presses."""
    text = update.message.text.strip()
    context.user_data["_redirect"] = text
    await show_main_menu(update, context, f"Switching to {text}...")
    return ConversationHandler.END


COMMON_FALLBACKS = [
    CommandHandler("start", restart_handler),
    MessageHandler(filters.Regex(r"^❌ Cancel$"), cancel_handler),
    MessageHandler(filters.Regex(MENU_PATTERN), menu_redirect),
]

# =============================================================================
# /START
# =============================================================================

async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await restart_handler(update, context)

# =============================================================================
# FEATURE 1: FILE ANALYSIS
# =============================================================================

async def fa_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "📊 *File Analysis*\n\n"
        "Upload a file (VCF, TXT, CSV, XLSX) and I'll analyze its contacts.\n\n"
        "👇 Send your file now.",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return FA_UPLOAD


async def fa_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Back button
    if update.message.text and update.message.text.strip() == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a file (VCF, TXT, CSV, XLSX).", reply_markup=back_cancel_kb())
        return FA_UPLOAD

    status_msg = await update.message.reply_text("⏳ Analyzing file...")
    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        contacts = parse_file(data, doc.file_name or "file.vcf")
        analysis = analyze_contacts(contacts)

        country_text = ""
        if analysis["countries"]:
            sorted_countries = sorted(analysis["countries"].items(), key=lambda x: -x[1])
            top = sorted_countries[:10]
            country_text = "\n".join(f"  • {c}: {n}" for c, n in top)
            if len(sorted_countries) > 10:
                country_text += f"\n  ...and {len(sorted_countries)-10} more countries"

        result = (
            f"📊 *File Analysis Result*\n\n"
            f"📁 File: `{doc.file_name}`\n\n"
            f"📞 Total contacts: *{analysis['total']}*\n"
            f"✅ Clean (unique): *{analysis['clean']}*\n"
            f"🔁 Duplicates: *{analysis['duplicate']}*\n"
            f"🗑 Junk (invalid): *{analysis['junk']}*\n"
        )
        if country_text:
            result += f"\n🌍 *Country Breakdown:*\n{country_text}\n"

        await status_msg.edit_text(result, parse_mode="Markdown")
    except Exception as e:
        logger.error(f"FA error: {e}")
        await status_msg.edit_text(f"❌ Error analyzing file: {e}")

    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 2: FILE CONVERTER
# =============================================================================

async def fc_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "🔄 *File Converter*\n\nUpload a file (VCF, TXT, CSV, XLSX) to convert.",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return FC_UPLOAD


async def fc_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.text and update.message.text.strip() == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a file.", reply_markup=back_cancel_kb())
        return FC_UPLOAD

    status = await update.message.reply_text("📥 Reading file...")
    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        contacts = parse_file(data, doc.file_name or "file.vcf")
        context.user_data["fc_contacts"] = contacts
        context.user_data["fc_original_name"] = clean_name(Path(doc.file_name or "file").stem) or "converted"
        await status.delete()
    except Exception as e:
        await status.edit_text(f"❌ Error: {e}")
        return ConversationHandler.END

    await update.message.reply_text(
        f"✅ Read *{len(contacts)}* contacts.\n\nChoose output format:",
        reply_markup=format_kb(),
        parse_mode="Markdown",
    )
    return FC_FORMAT


async def fc_format(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()

    if text == BTN_BACK:
        return await fc_entry(update, context)

    fmt_label_map = {"📄 TXT": "txt", "📇 VCF": "vcf", "📊 CSV": "csv", "📑 XLSX": "xlsx"}
    if text not in fmt_label_map:
        await update.message.reply_text("Please choose a format using the buttons.", reply_markup=format_kb())
        return FC_FORMAT

    fmt = fmt_label_map[text]
    contacts = context.user_data.get("fc_contacts", [])
    original_name = context.user_data.get("fc_original_name", "converted")

    status = await update.message.reply_text(f"⏳ Converting to {fmt.upper()}...")
    try:
        out_data = contacts_to_format(contacts, fmt)
        out_name = f"{original_name}.{fmt}"
        await update.message.reply_document(
            document=InputFile(io.BytesIO(out_data), filename=out_name),
            caption=f"✅ Converted {len(contacts)} contacts to {fmt.upper()}.",
        )
        await status.delete()
    except Exception as e:
        await status.edit_text(f"❌ Conversion error: {e}")
        return ConversationHandler.END

    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 3: QUICK VCF
# =============================================================================

async def qv_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["qv_contacts"] = []
    await update.message.reply_text(
        "⚡ *Quick VCF*\n\nEnter the VCF filename (without extension):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return QV_FILENAME


async def qv_filename(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    filename = clean_name(text) or "contacts"
    context.user_data["qv_filename"] = filename
    await update.message.reply_text("👤 Enter contact name:", reply_markup=back_cancel_kb())
    return QV_NAME


async def qv_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        return await qv_entry(update, context)

    context.user_data["qv_current_name"] = text
    await update.message.reply_text("📞 Enter phone number:", reply_markup=back_cancel_kb())
    return QV_PHONE


async def qv_phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text("👤 Enter contact name:", reply_markup=back_cancel_kb())
        return QV_NAME

    phone = clean_phone(text)
    if not phone:
        await update.message.reply_text(
            "❌ Invalid phone number (need 7+ digits). Try again:",
            reply_markup=back_cancel_kb(),
        )
        return QV_PHONE

    name = context.user_data.get("qv_current_name", "Contact")
    context.user_data["qv_contacts"].append((name, phone))
    count = len(context.user_data["qv_contacts"])

    await update.message.reply_text(
        f"✅ Added: *{name}* — `{phone}`\n📊 Total: {count} contact(s)\n\nAdd more?",
        reply_markup=more_finish_kb(),
        parse_mode="Markdown",
    )
    return QV_MORE


async def qv_more(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "➕ Add More":
        await update.message.reply_text("👤 Enter next contact name:", reply_markup=back_cancel_kb())
        return QV_NAME
    elif text == "✅ Finish":
        return await qv_finish(update, context)
    else:
        await update.message.reply_text("Please use the buttons.", reply_markup=more_finish_kb())
        return QV_MORE


async def qv_finish(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    contacts = context.user_data.get("qv_contacts", [])
    filename = context.user_data.get("qv_filename", "contacts")

    if not contacts:
        await update.message.reply_text("No contacts added.")
        await show_main_menu(update, context)
        return ConversationHandler.END

    vcf_data = contacts_to_vcf(contacts)
    out_name = f"{filename}.vcf"
    await update.message.reply_document(
        document=InputFile(io.BytesIO(vcf_data), filename=out_name),
        caption=f"✅ Quick VCF created with {len(contacts)} contact(s).",
    )
    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 4: VCF MAKER — COMPLETELY NEW 8-STEP FLOW (CHANGE 3, 4, 5)
# =============================================================================
# Always outputs .vcf files. No format selection step.
#
# Step 1: Upload source file (VM_UPLOAD)
# Step 2: VCF file base name (VM_BASENAME)
# Step 3: Contact base name (VM_CONTACT_NAME)
# Step 4: How many contacts per VCF file (VM_PER_FILE)
# Step 5: Contact numbering starts from? (VM_CONTACT_START)
# Step 6: VCF file numbering starts from? (VM_FILE_START)
# Step 7: Group tag name or Skip (VM_GROUP_NAME)
# Step 8: Group tag numbering starts from? (VM_GROUP_START) — only if group given
# Confirm → Generate (VM_CONFIRM)
# =============================================================================

async def vm_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["vm"] = {}
    await update.message.reply_text(
        "🛠 *VCF Maker* — Step 1/8\n\nUpload source file with phone numbers (TXT, CSV, XLSX, VCF):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return VM_UPLOAD


async def vm_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.text and update.message.text.strip() == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a file.", reply_markup=back_cancel_kb())
        return VM_UPLOAD

    status = await update.message.reply_text("📥 Reading numbers...")
    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        contacts = parse_file(data, doc.file_name or "file.txt")
        phones = [p for _, p in contacts]
        if not phones:
            await status.edit_text("❌ No valid phone numbers found in file.")
            return VM_UPLOAD
        context.user_data["vm"]["phones"] = phones
        await status.edit_text(f"✅ Found *{len(phones)}* numbers.", parse_mode="Markdown")
    except Exception as e:
        await status.edit_text(f"❌ Error: {e}")
        return VM_UPLOAD

    await update.message.reply_text(
        "Step 2/8 — Enter VCF file base name (e.g. `madara` → madara1.vcf, madara2.vcf):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return VM_BASENAME


async def vm_basename(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Step 1/8 — Upload source file with phone numbers:",
            reply_markup=back_cancel_kb(),
        )
        return VM_UPLOAD

    context.user_data["vm"]["basename"] = clean_name(text) or "contacts"
    await update.message.reply_text(
        "Step 3/8 — Enter contact base name (e.g. `Customer`):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return VM_CONTACT_NAME


async def vm_contact_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Step 2/8 — Enter VCF file base name:",
            reply_markup=back_cancel_kb(),
        )
        return VM_BASENAME

    context.user_data["vm"]["contact_name"] = text or "Contact"
    await update.message.reply_text(
        "Step 4/8 — How many contacts per VCF file? (e.g. `50`):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return VM_PER_FILE


async def vm_per_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Step 3/8 — Enter contact base name:",
            reply_markup=back_cancel_kb(),
        )
        return VM_CONTACT_NAME

    try:
        val = int(text)
        if val < 1:
            raise ValueError
        context.user_data["vm"]["per_file"] = val
    except ValueError:
        await update.message.reply_text("Please enter a valid positive number.", reply_markup=back_cancel_kb())
        return VM_PER_FILE

    await update.message.reply_text(
        "Step 5/8 — Contact numbering starts from? (e.g. `1`):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return VM_CONTACT_START


async def vm_contact_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Step 4/8 — How many contacts per VCF file?:",
            reply_markup=back_cancel_kb(),
        )
        return VM_PER_FILE

    try:
        context.user_data["vm"]["contact_start"] = int(text)
    except ValueError:
        await update.message.reply_text("Please enter a valid number.", reply_markup=back_cancel_kb())
        return VM_CONTACT_START

    await update.message.reply_text(
        "Step 6/8 — VCF file numbering starts from? (e.g. `1`):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return VM_FILE_START


async def vm_file_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Step 5/8 — Contact numbering starts from?:",
            reply_markup=back_cancel_kb(),
        )
        return VM_CONTACT_START

    try:
        context.user_data["vm"]["file_start"] = int(text)
    except ValueError:
        await update.message.reply_text("Please enter a valid number.", reply_markup=back_cancel_kb())
        return VM_FILE_START

    await update.message.reply_text(
        "Step 7/8 — Enter group tag name (e.g. `emonavuy`) or ⏭ Skip:",
        reply_markup=skip_cancel_kb(),
        parse_mode="Markdown",
    )
    return VM_GROUP_NAME


async def vm_group_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Step 6/8 — VCF file numbering starts from?:",
            reply_markup=back_cancel_kb(),
        )
        return VM_FILE_START

    if text == "⏭ Skip":
        context.user_data["vm"]["group_name"] = ""
        # Skip group start step — go straight to confirm
        return await vm_show_confirm(update, context)
    else:
        context.user_data["vm"]["group_name"] = text
        await update.message.reply_text(
            "Step 8/8 — Group tag numbering starts from? (e.g. `10`):",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return VM_GROUP_START


async def vm_group_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Step 7/8 — Enter group tag name or ⏭ Skip:",
            reply_markup=skip_cancel_kb(),
        )
        return VM_GROUP_NAME

    try:
        context.user_data["vm"]["group_start"] = int(text)
    except ValueError:
        await update.message.reply_text("Please enter a valid number.", reply_markup=back_cancel_kb())
        return VM_GROUP_START

    return await vm_show_confirm(update, context)


async def vm_show_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """CHANGE 5: Show confirm message with new format."""
    vm = context.user_data["vm"]
    phones = vm["phones"]
    basename = vm["basename"]
    per_file = vm["per_file"]
    contact_start = vm["contact_start"]
    file_start = vm["file_start"]
    group_name = vm.get("group_name", "")
    group_start = vm.get("group_start", 1)

    total = len(phones)
    total_files = (total + per_file - 1) // per_file
    file_end = file_start + total_files - 1

    if group_name:
        group_display = f"{group_name} (starts from {group_start})"
    else:
        group_display = "(none)"

    confirm_text = (
        f"✅ *Confirm VCF Maker Settings*\n\n"
        f"📞 Numbers: *{total}*\n"
        f"📁 Files: `{basename}{file_start}.vcf` ... `{basename}{file_end}.vcf`\n"
        f"👤 Contact name: `{vm['contact_name']}`\n"
        f"🔢 Contact numbering: starts from `{contact_start}`\n"
        f"📦 Per file: `{per_file}`\n"
        f"📂 File numbering: starts from `{file_start}`\n"
        f"🏷 Group: `{group_display}`\n"
        f"📂 Total files: *{total_files}*\n"
    )
    await update.message.reply_text(confirm_text, reply_markup=confirm_kb(), parse_mode="Markdown")
    return VM_CONFIRM


async def vm_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK or text == BTN_CANCEL:
        await show_main_menu(update, context)
        return ConversationHandler.END

    if text != "✅ Generate":
        await update.message.reply_text("Please press '✅ Generate' or '❌ Cancel'.", reply_markup=confirm_kb())
        return VM_CONFIRM

    await vm_generate(update, context)
    await show_main_menu(update, context)
    return ConversationHandler.END


async def vm_generate(update, context):
    """CHANGE 3: VCF Maker generation logic — always VCF, new naming scheme."""
    vm = context.user_data["vm"]
    phones = vm["phones"]
    basename = vm["basename"]
    contact_name = vm["contact_name"]
    per_file = vm["per_file"]
    contact_start = vm["contact_start"]
    file_start = vm["file_start"]
    group_name = vm.get("group_name", "")
    group_start = vm.get("group_start", 1)

    total = len(phones)
    file_idx = 0
    contact_num = contact_start

    status = await update.message.reply_text("⏳ Generating VCF files...")

    for chunk_start in range(0, total, per_file):
        chunk = phones[chunk_start:chunk_start + per_file]
        current_file_num = file_start + file_idx
        current_group_num = group_start + file_idx

        contacts = []
        for phone in chunk:
            if group_name:
                name = f"{contact_name} {contact_num} {group_name} {current_group_num}"
            else:
                name = f"{contact_name} {contact_num}"
            contacts.append((name, phone))
            contact_num += 1

        vcf_data = contacts_to_vcf(contacts)
        out_name = f"{basename}{current_file_num}.vcf"
        await update.message.reply_document(
            document=InputFile(io.BytesIO(vcf_data), filename=out_name),
            caption=f"📦 {out_name}: {len(contacts)} contacts",
        )
        file_idx += 1

    total_files = file_idx
    await status.edit_text(f"✅ Done! Generated {total_files} file(s) with {total} contacts total.")

# =============================================================================
# FEATURE 5: SPLIT FILE
# =============================================================================

async def sf_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "✂️ *Split File*\n\nUpload a file to split (VCF, TXT, CSV, XLSX):",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return SF_UPLOAD


async def sf_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.text and update.message.text.strip() == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a file.", reply_markup=back_cancel_kb())
        return SF_UPLOAD

    status = await update.message.reply_text("📥 Reading file...")
    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        contacts = parse_file(data, doc.file_name or "file.vcf")
        if not contacts:
            await status.edit_text("❌ No contacts found in file.")
            return SF_UPLOAD
        context.user_data["sf_contacts"] = contacts
        context.user_data["sf_ext"] = Path(doc.file_name or "file.vcf").suffix.lower().lstrip(".")
        raw_stem = Path(doc.file_name or "file").stem
        context.user_data["sf_stem"] = clean_name(raw_stem) or "file"
        await status.edit_text(f"✅ Found *{len(contacts)}* contacts.", parse_mode="Markdown")
    except Exception as e:
        await status.edit_text(f"❌ Error: {e}")
        return SF_UPLOAD

    await update.message.reply_text(
        "How many contacts per split file? (Enter a number):",
        reply_markup=back_cancel_kb(),
    )
    return SF_COUNT


async def sf_count(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        return await sf_entry(update, context)

    try:
        count = int(text)
        if count < 1:
            raise ValueError
    except ValueError:
        await update.message.reply_text("Please enter a valid positive number.", reply_markup=back_cancel_kb())
        return SF_COUNT

    contacts = context.user_data.get("sf_contacts", [])
    ext = context.user_data.get("sf_ext", "vcf")
    stem = context.user_data.get("sf_stem", "file")

    status = await update.message.reply_text("⏳ Splitting...")
    total = len(contacts)
    file_num = 1

    for chunk_start in range(0, total, count):
        chunk = contacts[chunk_start:chunk_start + count]
        out_data = contacts_to_format(chunk, ext)
        out_name = f"{stem}{file_num}.{ext}"
        await update.message.reply_document(
            document=InputFile(io.BytesIO(out_data), filename=out_name),
            caption=f"📦 Part {file_num}: {len(chunk)} contacts",
        )
        file_num += 1

    await status.edit_text(f"✅ Split into {file_num-1} file(s).")
    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 6: MERGE FILES
# =============================================================================

async def mf_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["mf_contacts"] = []
    context.user_data["mf_file_count"] = 0
    await update.message.reply_text(
        "🔗 *Merge Files*\n\n"
        "Upload files one by one.\n"
        "Press *✅ Done* when finished uploading.",
        reply_markup=done_cancel_kb(),
        parse_mode="Markdown",
    )
    return MF_UPLOAD


async def mf_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip() if update.message.text else ""

    if text == BTN_DONE:
        return await mf_done(update, context)
    if text == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text(
            "Please send a file, or press ✅ Done when finished.",
            reply_markup=done_cancel_kb(),
        )
        return MF_UPLOAD

    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        new_contacts = parse_file(data, doc.file_name or "file.vcf")
        context.user_data["mf_contacts"].extend(new_contacts)
        context.user_data["mf_file_count"] = context.user_data.get("mf_file_count", 0) + 1

        file_count = context.user_data["mf_file_count"]
        total_contacts = len(context.user_data["mf_contacts"])

        await update.message.reply_text(
            f"📊 Files: *{file_count}* | Contacts: *{total_contacts}*\n\nUpload more or press ✅ Done.",
            reply_markup=done_cancel_kb(),
            parse_mode="Markdown",
        )
    except Exception as e:
        logger.error(f"MF upload error: {e}")
        await update.message.reply_text(f"❌ Error reading file: {e}", reply_markup=done_cancel_kb())

    return MF_UPLOAD


async def mf_done(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    contacts = context.user_data.get("mf_contacts", [])
    if not contacts:
        await update.message.reply_text("❌ No contacts to merge.")
        await show_main_menu(update, context)
        return ConversationHandler.END

    status = await update.message.reply_text(f"⏳ Merging {len(contacts)} contacts...")
    vcf_data = contacts_to_vcf(contacts)
    await update.message.reply_document(
        document=InputFile(io.BytesIO(vcf_data), filename="merged.vcf"),
        caption=f"✅ Merged {len(contacts)} contacts from {context.user_data['mf_file_count']} files.",
    )
    await status.delete()
    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 7: FILE EDITOR
# =============================================================================

async def fe_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["fe_contacts"] = []
    context.user_data["fe_page"] = 0
    context.user_data["fe_ext"] = "vcf"
    await update.message.reply_text(
        "✏️ *File Editor*\n\nUpload a VCF file to edit:",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return FE_UPLOAD


async def fe_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.text and update.message.text.strip() == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a file.", reply_markup=back_cancel_kb())
        return FE_UPLOAD

    status = await update.message.reply_text("📥 Reading file...")
    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        contacts = parse_file(data, doc.file_name or "file.vcf")
        if not contacts:
            await status.edit_text("❌ No contacts found.")
            return FE_UPLOAD
        context.user_data["fe_contacts"] = list(contacts)
        context.user_data["fe_page"] = 0
        context.user_data["fe_ext"] = Path(doc.file_name or "file.vcf").suffix.lower().lstrip(".")
        raw_stem = Path(doc.file_name or "file").stem
        context.user_data["fe_stem"] = clean_name(raw_stem) or "file"
        await status.delete()
    except Exception as e:
        await status.edit_text(f"❌ Error: {e}")
        return ConversationHandler.END

    return await fe_show_page(update, context)


async def fe_show_page(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    contacts = context.user_data.get("fe_contacts", [])
    page = context.user_data.get("fe_page", 0)
    text, page, total_pages = paginate_contacts(contacts, page)
    context.user_data["fe_page"] = page
    text += f"\n\n_Use the buttons to navigate, edit, remove, add, or save._"
    await update.effective_message.reply_text(text, reply_markup=editor_kb(), parse_mode="Markdown")
    return FE_VIEW


async def fe_view(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    contacts = context.user_data.get("fe_contacts", [])

    if text == "◀️ Prev":
        context.user_data["fe_page"] = max(0, context.user_data.get("fe_page", 0) - 1)
        return await fe_show_page(update, context)

    elif text == "▶️ Next":
        per_page = 10
        total_pages = max(1, (len(contacts) + per_page - 1) // per_page)
        context.user_data["fe_page"] = min(total_pages - 1, context.user_data.get("fe_page", 0) + 1)
        return await fe_show_page(update, context)

    elif text == "✏️ Edit":
        total = len(contacts)
        await update.message.reply_text(
            f"✏️ Enter the contact number to edit (1–{total}):",
            reply_markup=back_cancel_kb(),
        )
        return FE_EDIT_SELECT

    elif text == "🗑 Remove":
        total = len(contacts)
        await update.message.reply_text(
            f"🗑 Enter the contact number to remove (1–{total}):",
            reply_markup=back_cancel_kb(),
        )
        return FE_REMOVE_IDX

    elif text == "➕ Add":
        await update.message.reply_text("➕ Enter new contact name:", reply_markup=back_cancel_kb())
        return FE_ADD_NAME

    elif text == "💾 Save":
        return await fe_save(update, context)

    elif text == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    else:
        await update.message.reply_text("Use the keyboard buttons.", reply_markup=editor_kb())
        return FE_VIEW


async def fe_edit_select(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        return await fe_show_page(update, context)

    try:
        idx = int(text) - 1
        contacts = context.user_data["fe_contacts"]
        if idx < 0 or idx >= len(contacts):
            raise IndexError
        context.user_data["fe_edit_idx"] = idx
        name, phone = contacts[idx]
        await update.message.reply_text(
            f"Editing: *{name}* — `{phone}`\n\nEnter new name (or `-` to keep):",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return FE_EDIT_NAME
    except (ValueError, IndexError):
        await update.message.reply_text("Invalid number. Try again:", reply_markup=back_cancel_kb())
        return FE_EDIT_SELECT


async def fe_edit_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        total = len(context.user_data.get("fe_contacts", []))
        await update.message.reply_text(
            f"Enter the contact number to edit (1–{total}):",
            reply_markup=back_cancel_kb(),
        )
        return FE_EDIT_SELECT

    idx = context.user_data.get("fe_edit_idx", 0)
    contacts = context.user_data["fe_contacts"]
    if text != "-":
        contacts[idx] = (text, contacts[idx][1])
    context.user_data["fe_contacts"] = contacts

    await update.message.reply_text(
        "Enter new phone number (or `-` to keep):",
        reply_markup=back_cancel_kb(),
    )
    return FE_EDIT_PHONE


async def fe_edit_phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        idx = context.user_data.get("fe_edit_idx", 0)
        contacts = context.user_data["fe_contacts"]
        name, phone = contacts[idx]
        await update.message.reply_text(
            f"Enter new name for *{name}* (or `-` to keep):",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return FE_EDIT_NAME

    idx = context.user_data.get("fe_edit_idx", 0)
    contacts = context.user_data["fe_contacts"]
    if text != "-":
        phone = clean_phone(text)
        if not phone:
            await update.message.reply_text("❌ Invalid phone. Try again:", reply_markup=back_cancel_kb())
            return FE_EDIT_PHONE
        contacts[idx] = (contacts[idx][0], phone)
    context.user_data["fe_contacts"] = contacts
    name, phone = contacts[idx]
    await update.message.reply_text(f"✅ Updated: *{name}* — `{phone}`", parse_mode="Markdown")
    return await fe_show_page(update, context)


async def fe_remove_idx(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        return await fe_show_page(update, context)

    try:
        idx = int(text) - 1
        contacts = context.user_data["fe_contacts"]
        if idx < 0 or idx >= len(contacts):
            raise IndexError
        removed = contacts.pop(idx)
        context.user_data["fe_contacts"] = contacts
        per_page = 10
        total_pages = max(1, (len(contacts) + per_page - 1) // per_page)
        if context.user_data.get("fe_page", 0) >= total_pages:
            context.user_data["fe_page"] = max(0, total_pages - 1)
        await update.message.reply_text(
            f"✅ Removed: *{removed[0]}* — `{removed[1]}`", parse_mode="Markdown"
        )
        return await fe_show_page(update, context)
    except (ValueError, IndexError):
        await update.message.reply_text("Invalid number. Try again:", reply_markup=back_cancel_kb())
        return FE_REMOVE_IDX


async def fe_add_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        return await fe_show_page(update, context)

    context.user_data["fe_new_name"] = text
    await update.message.reply_text("Enter new contact phone number:", reply_markup=back_cancel_kb())
    return FE_ADD_PHONE


async def fe_add_phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text("➕ Enter new contact name:", reply_markup=back_cancel_kb())
        return FE_ADD_NAME

    phone = clean_phone(text)
    if not phone:
        await update.message.reply_text("❌ Invalid phone. Try again:", reply_markup=back_cancel_kb())
        return FE_ADD_PHONE

    name = context.user_data.get("fe_new_name", "Contact")
    context.user_data["fe_contacts"].append((name, phone))
    await update.message.reply_text(f"✅ Added: *{name}* — `{phone}`", parse_mode="Markdown")
    return await fe_show_page(update, context)


async def fe_save(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    contacts = context.user_data.get("fe_contacts", [])
    ext = context.user_data.get("fe_ext", "vcf")
    stem = context.user_data.get("fe_stem", "file")

    out_data = contacts_to_format(contacts, ext)
    out_name = f"{stem}edited.{ext}"
    await update.effective_message.reply_document(
        document=InputFile(io.BytesIO(out_data), filename=out_name),
        caption=f"✅ Saved {len(contacts)} contacts.",
    )
    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 8: LIST MAKER (OCR with graceful fallback)
# =============================================================================

def ocr_image(img_bytes: bytes) -> str:
    if not TESSERACT_AVAILABLE:
        raise RuntimeError("pytesseract not installed")
    if not img_bytes or len(img_bytes) == 0:
        raise ValueError("Empty image data")
    try:
        image = Image.open(io.BytesIO(img_bytes))
        text = pytesseract.image_to_string(image, lang="eng")
        return text
    except Exception as e:
        logger.error(f"OCR error: {e}")
        return ""


def parse_group_list(ocr_text: str) -> List[Tuple[str, str]]:
    results = []
    tier1 = re.findall(
        r'([A-Za-z0-9 _\-]+?)\s*[:\-]\s*(\d+)\s*pending',
        ocr_text, re.IGNORECASE
    )
    if tier1:
        return [(n.strip(), c.strip()) for n, c in tier1]

    tier2 = re.findall(
        r'^([A-Za-z][A-Za-z0-9 _\-]{2,}?)\s+(\d{1,5})\s*$',
        ocr_text, re.MULTILINE
    )
    if tier2:
        return [(n.strip(), c.strip()) for n, c in tier2]

    tier3 = re.findall(r'([A-Za-z][A-Za-z0-9 _\-]{2,})\D+(\d{1,5})', ocr_text)
    if tier3:
        return [(n.strip(), c.strip()) for n, c in tier3]

    for line in ocr_text.splitlines():
        line = line.strip()
        if line and re.search(r'[A-Za-z]', line) and re.search(r'\d', line):
            results.append((line, "?"))

    return results


async def lm_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["lm_images"] = []

    if not TESSERACT_AVAILABLE:
        await update.message.reply_text(
            "📋 *List Maker*\n\n"
            "⚠️ *Note:* Tesseract OCR is not installed on this server.\n"
            "Install `tesseract-ocr` to use this feature.\n\n"
            "```\nUbuntu: sudo apt-get install tesseract-ocr\n"
            "macOS: brew install tesseract\n```\n\n"
            "You can still upload screenshots and I'll try to process them, "
            "but OCR will not work without Tesseract.",
            reply_markup=done_cancel_kb(),
            parse_mode="Markdown",
        )
    else:
        await update.message.reply_text(
            "📋 *List Maker*\n\n"
            "Upload screenshots one by one.\n"
            "Press *✅ Done* when finished.",
            reply_markup=done_cancel_kb(),
            parse_mode="Markdown",
        )
    return LM_UPLOAD


async def lm_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip() if update.message.text else ""

    if text == BTN_DONE:
        return await lm_done(update, context)
    if text == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    photo = update.message.photo
    doc = update.message.document

    img_bytes = None
    try:
        if photo:
            tg_file = await context.bot.get_file(photo[-1].file_id)
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            buf.seek(0)
            img_bytes = buf.read()
            if not img_bytes or len(img_bytes) == 0:
                await update.message.reply_text("❌ Failed to download image. Try again.", reply_markup=done_cancel_kb())
                return LM_UPLOAD
        elif doc and doc.mime_type and doc.mime_type.startswith("image/"):
            tg_file = await context.bot.get_file(doc.file_id)
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            buf.seek(0)
            img_bytes = buf.read()
            if not img_bytes or len(img_bytes) == 0:
                await update.message.reply_text("❌ Failed to download image. Try again.", reply_markup=done_cancel_kb())
                return LM_UPLOAD
        else:
            await update.message.reply_text(
                "Please send an image/screenshot, or press ✅ Done.",
                reply_markup=done_cancel_kb(),
            )
            return LM_UPLOAD
    except Exception as e:
        logger.error(f"LM image download error: {e}")
        await update.message.reply_text(f"❌ Error receiving image: {e}", reply_markup=done_cancel_kb())
        return LM_UPLOAD

    context.user_data["lm_images"].append(img_bytes)
    count = len(context.user_data["lm_images"])
    await update.message.reply_text(
        f"📸 *{count}* screenshot(s) received. Upload more or press ✅ Done.",
        reply_markup=done_cancel_kb(),
        parse_mode="Markdown",
    )
    return LM_UPLOAD


async def lm_done(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    images = context.user_data.get("lm_images", [])
    if not images:
        await update.message.reply_text("❌ No screenshots uploaded.")
        await show_main_menu(update, context)
        return ConversationHandler.END

    if not TESSERACT_AVAILABLE:
        await update.message.reply_text(
            "❌ Cannot process OCR: Install tesseract-ocr for this feature.\n\n"
            "```\nUbuntu: sudo apt-get install tesseract-ocr\n"
            "macOS: brew install tesseract\n```",
            parse_mode="Markdown",
        )
        await show_main_menu(update, context)
        return ConversationHandler.END

    status = await update.message.reply_text(f"⏳ Processing {len(images)} screenshot(s) with OCR...")

    try:
        all_groups = []
        raw_texts = []
        for img_bytes in images:
            try:
                text = ocr_image(img_bytes)
                raw_texts.append(text)
                groups = parse_group_list(text)
                all_groups.extend(groups)
            except RuntimeError:
                await status.edit_text("❌ Install tesseract-ocr for this feature.")
                await show_main_menu(update, context)
                return ConversationHandler.END
            except Exception as e:
                logger.error(f"OCR processing error: {e}")
                raw_texts.append(f"[OCR Error: {e}]")

        if not all_groups:
            raw_combined = "\n---\n".join(raw_texts)
            await status.edit_text(
                "⚠️ Could not parse group names. Raw OCR text:\n\n" + raw_combined[:4000]
            )
            txt_data = raw_combined.encode("utf-8")
            await update.message.reply_document(
                document=InputFile(io.BytesIO(txt_data), filename="rawocr.txt"),
                caption="Raw OCR output",
            )
            await show_main_menu(update, context)
            return ConversationHandler.END

        lines = ["📋 *Group List:*\n"]
        txt_lines = ["Group List:\n"]
        for i, (name, count) in enumerate(all_groups, 1):
            lines.append(f"{i}. *{name}* — {count}")
            txt_lines.append(f"{i}. {name} — {count}")

        formatted = "\n".join(lines)
        txt_content = "\n".join(txt_lines)

        await status.delete()

        if len(formatted) > 4000:
            for chunk_start in range(0, len(lines), 50):
                chunk = "\n".join(lines[chunk_start:chunk_start + 50])
                await update.message.reply_text(chunk, parse_mode="Markdown")
        else:
            await update.message.reply_text(formatted, parse_mode="Markdown")

        await update.message.reply_document(
            document=InputFile(io.BytesIO(txt_content.encode("utf-8")), filename="grouplist.txt"),
            caption=f"✅ Extracted {len(all_groups)} group entries.",
        )
    except Exception as e:
        logger.error(f"LM done error: {e}")
        await status.edit_text(f"❌ Error during OCR processing: {e}")

    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 9: RENAME FILE
# =============================================================================

async def rf_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "📝 *Rename File*\n\nUpload the file you want to rename:",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return RF_UPLOAD


async def rf_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.text and update.message.text.strip() == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a file.", reply_markup=back_cancel_kb())
        return RF_UPLOAD

    status = await update.message.reply_text("📥 Reading file...")
    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        context.user_data["rf_data"] = data
        context.user_data["rf_ext"] = Path(doc.file_name or "file").suffix
        await status.delete()
    except Exception as e:
        await status.edit_text(f"❌ Error: {e}")
        return ConversationHandler.END

    await update.message.reply_text(
        "Enter new filename (without extension):",
        reply_markup=back_cancel_kb(),
    )
    return RF_NAME


async def rf_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        return await rf_entry(update, context)

    new_name = clean_name(text) or "renamed"
    ext = context.user_data.get("rf_ext", ".vcf")
    data = context.user_data.get("rf_data", b"")
    out_name = f"{new_name}{ext}"
    await update.message.reply_document(
        document=InputFile(io.BytesIO(data), filename=out_name),
        caption=f"✅ File renamed to `{out_name}`",
        parse_mode="Markdown",
    )
    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 10: RENAME CONTACT (BULK)
# =============================================================================

async def rc_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "👤 *Rename Contact*\n\nUpload a VCF file:",
        reply_markup=back_cancel_kb(),
        parse_mode="Markdown",
    )
    return RC_UPLOAD


async def rc_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.text and update.message.text.strip() == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a VCF file.", reply_markup=back_cancel_kb())
        return RC_UPLOAD

    status = await update.message.reply_text("📥 Reading VCF...")
    try:
        file = await context.bot.get_file(doc.file_id)
        data = bytes(await file.download_as_bytearray())
        contacts = parse_vcf_contacts(data)
        if not contacts:
            await status.edit_text("❌ No contacts found.")
            return RC_UPLOAD
        context.user_data["rc_contacts"] = list(contacts)
        raw_stem = Path(doc.file_name or "file").stem
        context.user_data["rc_stem"] = clean_name(raw_stem) or "file"
        await status.delete()
    except Exception as e:
        await status.edit_text(f"❌ Error: {e}")
        return ConversationHandler.END

    await update.message.reply_text(
        f"✅ Loaded *{len(context.user_data['rc_contacts'])}* contacts.\n\nRename ALL or SINGLE contact?",
        reply_markup=rename_mode_kb(),
        parse_mode="Markdown",
    )
    return RC_MODE


async def rc_mode(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()

    if text == BTN_BACK or text == BTN_CANCEL:
        await show_main_menu(update, context)
        return ConversationHandler.END

    if text == "👥 Rename ALL":
        await update.message.reply_text(
            "Enter new base name for ALL contacts (e.g. `Customer`):\n"
            "_All contacts will be renamed Customer 1, Customer 2..._",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return RC_ALL_NAME

    elif text == "👤 Rename SINGLE":
        contacts = context.user_data.get("rc_contacts", [])
        lines = ["Choose a contact to rename:\n"]
        for i, (name, phone) in enumerate(contacts[:50], 1):
            lines.append(f"`{i}.` {name} — `{phone}`")
        if len(contacts) > 50:
            lines.append(f"\n_(showing first 50 of {len(contacts)})_")
        lines.append("\nEnter contact number:")
        await update.message.reply_text("\n".join(lines), reply_markup=back_cancel_kb(), parse_mode="Markdown")
        return RC_SINGLE_SELECT

    else:
        await update.message.reply_text("Please use the buttons.", reply_markup=rename_mode_kb())
        return RC_MODE


async def rc_all_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Rename ALL or SINGLE contact?",
            reply_markup=rename_mode_kb(),
        )
        return RC_MODE

    base_name = text or "Contact"
    contacts = context.user_data.get("rc_contacts", [])
    renamed = [(f"{base_name} {i}", phone) for i, (_, phone) in enumerate(contacts, 1)]
    context.user_data["rc_contacts"] = renamed

    stem = context.user_data.get("rc_stem", "file")
    vcf_data = contacts_to_vcf(renamed)
    out_name = f"{stem}renamed.vcf"
    await update.message.reply_document(
        document=InputFile(io.BytesIO(vcf_data), filename=out_name),
        caption=f"✅ Renamed all {len(renamed)} contacts as '{base_name} N'.",
    )
    await show_main_menu(update, context)
    return ConversationHandler.END


async def rc_single_select(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        await update.message.reply_text(
            "Rename ALL or SINGLE contact?",
            reply_markup=rename_mode_kb(),
        )
        return RC_MODE

    try:
        idx = int(text) - 1
        contacts = context.user_data["rc_contacts"]
        if idx < 0 or idx >= len(contacts):
            raise IndexError
        context.user_data["rc_edit_idx"] = idx
        name, phone = contacts[idx]
        await update.message.reply_text(
            f"Renaming: *{name}* — `{phone}`\n\nEnter new name:",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return RC_SINGLE_NAME
    except (ValueError, IndexError):
        await update.message.reply_text("Invalid number. Try again:", reply_markup=back_cancel_kb())
        return RC_SINGLE_SELECT


async def rc_single_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == BTN_BACK:
        contacts = context.user_data.get("rc_contacts", [])
        lines = ["Choose a contact to rename:\n"]
        for i, (name, phone) in enumerate(contacts[:50], 1):
            lines.append(f"`{i}.` {name} — `{phone}`")
        lines.append("\nEnter contact number:")
        await update.message.reply_text("\n".join(lines), reply_markup=back_cancel_kb(), parse_mode="Markdown")
        return RC_SINGLE_SELECT

    new_name = text
    idx = context.user_data.get("rc_edit_idx", 0)
    contacts = context.user_data["rc_contacts"]
    contacts[idx] = (new_name, contacts[idx][1])
    context.user_data["rc_contacts"] = contacts

    stem = context.user_data.get("rc_stem", "file")
    vcf_data = contacts_to_vcf(contacts)
    out_name = f"{stem}renamed.vcf"
    await update.message.reply_document(
        document=InputFile(io.BytesIO(vcf_data), filename=out_name),
        caption=f"✅ Renamed contact to '{new_name}'.",
    )
    await show_main_menu(update, context)
    return ConversationHandler.END

# =============================================================================
# FEATURE 11: SETTINGS (Reply Keyboard)
# =============================================================================

async def st_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    settings = get_settings(context)
    await update.message.reply_text(
        "⚙️ *Settings*\n\nTap a setting button to change it:",
        reply_markup=settings_kb(settings),
        parse_mode="Markdown",
    )
    return ST_MAIN


async def st_main(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    settings = get_settings(context)

    if text == BTN_BACK:
        await show_main_menu(update, context)
        return ConversationHandler.END

    if text.startswith("📄 Format:"):
        # Cycle through formats
        formats = ["vcf", "txt", "csv", "xlsx"]
        cur = settings["default_format"]
        next_fmt = formats[(formats.index(cur) + 1) % len(formats)] if cur in formats else "vcf"
        settings["default_format"] = next_fmt
        await update.message.reply_text(
            f"✅ Default format set to *{next_fmt.upper()}*.\n\nTap again to change:",
            reply_markup=settings_kb(settings),
            parse_mode="Markdown",
        )
        return ST_MAIN

    elif text.startswith("📦 Per File:"):
        await update.message.reply_text(
            "Enter new contacts-per-file count (e.g. `500`):",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return ST_SET_PER_FILE

    elif text.startswith("👤 Contact Name:"):
        await update.message.reply_text(
            "Enter new default contact name (e.g. `Customer`):",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return ST_SET_CONTACT_NAME

    elif text.startswith("🔢 Start Number:"):
        await update.message.reply_text(
            "Enter new start number (e.g. `1`):",
            reply_markup=back_cancel_kb(),
            parse_mode="Markdown",
        )
        return ST_SET_START_NUM

    else:
        await update.message.reply_text(
            "Use the buttons to change settings.",
            reply_markup=settings_kb(settings),
        )
        return ST_MAIN


async def st_set_per_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    settings = get_settings(context)

    if text == BTN_BACK:
        await update.message.reply_text(
            "⚙️ *Settings*\n\nTap a setting button to change it:",
            reply_markup=settings_kb(settings),
            parse_mode="Markdown",
        )
        return ST_MAIN

    try:
        val = int(text)
        if val < 1:
            raise ValueError
        settings["per_file"] = val
        await update.message.reply_text(
            f"✅ Per-file count set to *{val}*.",
            reply_markup=settings_kb(settings),
            parse_mode="Markdown",
        )
    except ValueError:
        await update.message.reply_text("❌ Invalid. Enter a positive number.", reply_markup=back_cancel_kb())
        return ST_SET_PER_FILE

    return ST_MAIN


async def st_set_contact_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    settings = get_settings(context)

    if text == BTN_BACK:
        await update.message.reply_text(
            "⚙️ *Settings*",
            reply_markup=settings_kb(settings),
        )
        return ST_MAIN

    if not text:
        await update.message.reply_text("❌ Name cannot be empty.", reply_markup=back_cancel_kb())
        return ST_SET_CONTACT_NAME

    settings["contact_name"] = text
    await update.message.reply_text(
        f"✅ Contact name set to *{text}*.",
        reply_markup=settings_kb(settings),
        parse_mode="Markdown",
    )
    return ST_MAIN


async def st_set_start_num(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    settings = get_settings(context)

    if text == BTN_BACK:
        await update.message.reply_text(
            "⚙️ *Settings*",
            reply_markup=settings_kb(settings),
        )
        return ST_MAIN

    try:
        val = int(text)
        settings["start_number"] = val
        await update.message.reply_text(
            f"✅ Start number set to *{val}*.",
            reply_markup=settings_kb(settings),
            parse_mode="Markdown",
        )
    except ValueError:
        await update.message.reply_text("❌ Invalid. Enter a number.", reply_markup=back_cancel_kb())
        return ST_SET_START_NUM

    return ST_MAIN

# =============================================================================
# FEATURE 12: RESET
# =============================================================================

async def rs_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "🔄 *Reset Settings*\n\nAre you sure you want to reset all settings to defaults?",
        reply_markup=yes_no_kb(),
        parse_mode="Markdown",
    )
    return RS_CONFIRM


async def rs_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "✅ Yes, Reset":
        context.user_data["settings"] = DEFAULT_SETTINGS.copy()
        await update.message.reply_text("✅ Settings have been reset to defaults.")
        await show_main_menu(update, context)
    else:
        await show_main_menu(update, context, "❌ Reset cancelled.")
    return ConversationHandler.END

# =============================================================================
# FEATURE 13: HELP
# =============================================================================

async def help_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (
        "❓ *VCF Contact Bot — Help Guide*\n\n"
        "*📊 File Analysis*\n"
        "Upload VCF/TXT/CSV/XLSX → get total, clean, duplicate, junk counts + country breakdown.\n\n"
        "*🔄 File Converter*\n"
        "Upload any supported file → choose output format → get converted file.\n\n"
        "*⚡ Quick VCF*\n"
        "Create a VCF from scratch. Enter filename → add contacts (name + phone) → Done.\n\n"
        "*🛠 VCF Maker*\n"
        "8-step wizard: upload numbers → base name → contact name → per-file → contact start → "
        "file start → group tag (optional) → group start (if group given) → Confirm & Generate.\n"
        "Always outputs .vcf files. Contact numbering is continuous across all files. "
        "Group tag number increments per file.\n\n"
        "*✂️ Split File*\n"
        "Upload file → enter count per file → get split files.\n\n"
        "*🔗 Merge Files*\n"
        "Upload multiple files → Done → get single merged.vcf.\n\n"
        "*✏️ File Editor*\n"
        "Upload file → browse/edit/remove/add contacts → Save.\n\n"
        "*📋 List Maker*\n"
        "Upload group screenshots → Done → OCR extracts group names + pending counts → get list + TXT.\n"
        "Requires tesseract-ocr to be installed.\n\n"
        "*📝 Rename File*\n"
        "Upload file → enter new name → get renamed file.\n\n"
        "*👤 Rename Contact*\n"
        "Upload VCF → rename ALL (base name) or SINGLE (pick one) → get updated VCF.\n\n"
        "*⚙️ Settings*\n"
        "Tap buttons to set: default format, per-file count, contact name, start number.\n\n"
        "*🔄 Reset*\n"
        "Resets all settings to defaults.\n\n"
        "📞 *Phone Number Formats Accepted:*\n"
        "+919876543210, 919876543210, 9876543210, +91 98765 43210, 98765-43210\n"
        "_Any format with 7+ digits is accepted. All numbers get a + prefix._\n\n"
        "📁 *Filename Convention:*\n"
        "Filenames allow letters, digits, hyphens (-) and underscores (_).\n"
        "e.g. my_file-v2.vcf, madara1.vcf, madara2.vcf"
    )
    await update.message.reply_text(text, reply_markup=main_menu_kb(), parse_mode="Markdown")
    return ConversationHandler.END

# =============================================================================
# CONVERSATION HANDLERS
# =============================================================================

def build_fa_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_FA)}$"), fa_entry)],
        states={
            FA_UPLOAD: [
                MessageHandler(filters.Regex(r"^🔙 Back$"), fa_upload),
                MessageHandler(filters.Document.ALL, fa_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, fa_upload),
            ],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_fc_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_FC)}$"), fc_entry)],
        states={
            FC_UPLOAD: [
                MessageHandler(filters.Document.ALL, fc_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, fc_upload),
            ],
            FC_FORMAT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, fc_format),
            ],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_qv_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_QV)}$"), qv_entry)],
        states={
            QV_FILENAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, qv_filename)],
            QV_NAME:     [MessageHandler(filters.TEXT & ~filters.COMMAND, qv_name)],
            QV_PHONE:    [MessageHandler(filters.TEXT & ~filters.COMMAND, qv_phone)],
            QV_MORE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, qv_more)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_vm_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_VM)}$"), vm_entry)],
        states={
            VM_UPLOAD:        [
                MessageHandler(filters.Document.ALL, vm_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, vm_upload),
            ],
            VM_BASENAME:      [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_basename)],
            VM_CONTACT_NAME:  [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_contact_name)],
            VM_PER_FILE:      [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_per_file)],
            VM_CONTACT_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_contact_start)],
            VM_FILE_START:    [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_file_start)],
            VM_GROUP_NAME:    [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_group_name)],
            VM_GROUP_START:   [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_group_start)],
            VM_CONFIRM:       [MessageHandler(filters.TEXT & ~filters.COMMAND, vm_confirm)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_sf_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_SF)}$"), sf_entry)],
        states={
            SF_UPLOAD: [
                MessageHandler(filters.Document.ALL, sf_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, sf_upload),
            ],
            SF_COUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, sf_count)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_mf_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_MF)}$"), mf_entry)],
        states={
            MF_UPLOAD: [
                MessageHandler(filters.Document.ALL, mf_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, mf_upload),
            ],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_fe_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_FE)}$"), fe_entry)],
        states={
            FE_UPLOAD: [
                MessageHandler(filters.Document.ALL, fe_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, fe_upload),
            ],
            FE_VIEW: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, fe_view),
            ],
            FE_EDIT_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, fe_edit_select)],
            FE_EDIT_NAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, fe_edit_name)],
            FE_EDIT_PHONE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, fe_edit_phone)],
            FE_REMOVE_IDX:  [MessageHandler(filters.TEXT & ~filters.COMMAND, fe_remove_idx)],
            FE_ADD_NAME:    [MessageHandler(filters.TEXT & ~filters.COMMAND, fe_add_name)],
            FE_ADD_PHONE:   [MessageHandler(filters.TEXT & ~filters.COMMAND, fe_add_phone)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_lm_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_LM)}$"), lm_entry)],
        states={
            LM_UPLOAD: [
                MessageHandler(filters.PHOTO | filters.Document.IMAGE, lm_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, lm_upload),
            ],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_rf_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_RF)}$"), rf_entry)],
        states={
            RF_UPLOAD: [
                MessageHandler(filters.Document.ALL, rf_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, rf_upload),
            ],
            RF_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, rf_name)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_rc_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_RC)}$"), rc_entry)],
        states={
            RC_UPLOAD:        [
                MessageHandler(filters.Document.ALL, rc_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, rc_upload),
            ],
            RC_MODE:          [MessageHandler(filters.TEXT & ~filters.COMMAND, rc_mode)],
            RC_ALL_NAME:      [MessageHandler(filters.TEXT & ~filters.COMMAND, rc_all_name)],
            RC_SINGLE_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, rc_single_select)],
            RC_SINGLE_NAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, rc_single_name)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_st_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_ST)}$"), st_entry)],
        states={
            ST_MAIN:             [MessageHandler(filters.TEXT & ~filters.COMMAND, st_main)],
            ST_SET_PER_FILE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, st_set_per_file)],
            ST_SET_CONTACT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, st_set_contact_name)],
            ST_SET_START_NUM:    [MessageHandler(filters.TEXT & ~filters.COMMAND, st_set_start_num)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_rs_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_RS)}$"), rs_entry)],
        states={
            RS_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, rs_confirm)],
        },
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )


def build_help_conv() -> ConversationHandler:
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_HELP)}$"), help_entry)],
        states={},
        fallbacks=COMMON_FALLBACKS,
        allow_reentry=True,
        per_message=False,
    )

# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    if BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        raise RuntimeError("Set BOT_TOKEN environment variable or edit BOT_TOKEN in bot.py")

    app = Application.builder().token(BOT_TOKEN).build()

    # /start command
    app.add_handler(CommandHandler("start", start_cmd))

    # All feature conversation handlers
    app.add_handler(build_fa_conv())
    app.add_handler(build_fc_conv())
    app.add_handler(build_qv_conv())
    app.add_handler(build_vm_conv())
    app.add_handler(build_sf_conv())
    app.add_handler(build_mf_conv())
    app.add_handler(build_fe_conv())
    app.add_handler(build_lm_conv())
    app.add_handler(build_rf_conv())
    app.add_handler(build_rc_conv())
    app.add_handler(build_st_conv())
    app.add_handler(build_rs_conv())
    app.add_handler(build_help_conv())

    logger.info("Bot is running. Press Ctrl+C to stop.")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
